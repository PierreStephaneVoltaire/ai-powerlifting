"""Deterministic pre-classification and extraction for import files.

Identifies if a file is a Template (date-free) or a Session Import (dated)
before falling back to AI classification.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
from datetime import datetime, date
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)

def file_hash(file_bytes: bytes) -> str:
    """Return a short sha256 hash prefix for the file."""
    h = hashlib.sha256(file_bytes).hexdigest()
    return f"sha256:{h[:16]}"

def extract_xlsx(file_bytes: bytes) -> tuple[list[dict[str, Any]], str]:
    """Extract rows from all non-empty sheets of an XLSX file.

    Each row gets a ``_sheet`` key so downstream consumers can distinguish
    which sheet it came from.  Returns (rows, comma-separated sheet names).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    all_rows: list[dict[str, Any]] = []
    sheet_names: list[str] = []

    for sname in wb.sheetnames:
        sheet = wb[sname]
        if sheet.max_row is None or sheet.max_row < 2:
            continue

        first_row = list(sheet.iter_rows(min_row=1, max_row=1))
        if not first_row:
            continue
        headers = [
            str(cell.value).strip() if cell.value else f"col_{i}"
            for i, cell in enumerate(first_row[0])
        ]

        sheet_has_data = False
        for row_cells in sheet.iter_rows(min_row=2):
            row_dict: dict[str, Any] = {"_sheet": sname}
            has_data = False
            for i, cell in enumerate(row_cells):
                if i < len(headers):
                    val = cell.value
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()
                    row_dict[headers[i]] = val
                    if val is not None:
                        has_data = True
            if has_data:
                all_rows.append(row_dict)
                sheet_has_data = True

        if sheet_has_data:
            sheet_names.append(sname)

    return all_rows, ", ".join(sheet_names) if sheet_names else "empty"

def extract_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    """Extract rows from a CSV file."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader if any(row.values())]

def preclassify_rows(rows: list[dict[str, Any]]) -> Optional[str]:
    """Identify 'template' vs 'session_import' via heuristics.
    
    Returns 'template', 'session_import', or None if ambiguous.
    """
    if not rows:
        return None
        
    date_cols = [k for k in rows[0].keys() if "date" in k.lower()]
    has_actual_dates = False
    for row in rows[:20]:
        for col in date_cols:
            val = row.get(col)
            if val and (isinstance(val, (datetime, date)) or (isinstance(val, str) and "-" in val and len(val) >= 8)):
                has_actual_dates = True
                break
        if has_actual_dates: break

    if has_actual_dates:
        return "session_import"

    week_cols = [k for k in rows[0].keys() if "week" in k.lower()]
    if week_cols:
        return "template"

    load_cols = [k for k in rows[0].keys() if any(x in k.lower() for x in ["rpe", "percentage", "%", "target"])]
    if load_cols and not has_actual_dates:
        return "template"

    return None
