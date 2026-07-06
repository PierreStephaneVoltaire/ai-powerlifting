import asyncio
import base64
import csv
import hashlib
import io
import json
import os
from datetime import date, datetime
from typing import Any, Optional

_program_store: Optional[object] = None
_glossary_store: Optional[object] = None
_import_store: Optional[object] = None


def _run_async(coro):
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None
    if loop and loop.is_running():
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor() as pool:
            return pool.submit(asyncio.run, coro).result()
    return asyncio.run(coro)


def _get_program_store():
    global _program_store
    if _program_store is None:
        from program_store import ProgramStore
        _program_store = ProgramStore(
            table_name=os.environ.get("IF_HEALTH_TABLE_NAME", "if-health"),
            pk=os.environ.get("HEALTH_PROGRAM_PK", "operator"),
            region=os.environ.get("AWS_REGION", "ca-central-1"),
        )
    return _program_store


def _get_glossary_store():
    global _glossary_store
    if _glossary_store is None:
        from glossary_store import GlossaryStore
        _glossary_store = GlossaryStore(
            table_name=os.environ.get("IF_HEALTH_TABLE_NAME", "if-health"),
            pk=os.environ.get("HEALTH_PROGRAM_PK", "operator"),
            region=os.environ.get("AWS_REGION", "ca-central-1"),
        )
    return _glossary_store


def _get_import_store():
    global _import_store
    if _import_store is None:
        from import_store import ImportStore
        _import_store = ImportStore(
            table_name=os.environ.get("IF_HEALTH_TABLE_NAME", "if-health"),
            pk=os.environ.get("HEALTH_PROGRAM_PK", "operator"),
            region=os.environ.get("AWS_REGION", "ca-central-1"),
        )
    return _import_store


def _file_hash(file_bytes: bytes) -> str:
    h = hashlib.sha256(file_bytes).hexdigest()
    return f"sha256:{h[:16]}"


def _extract_xlsx(file_bytes: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_rows: list[dict[str, Any]] = []
    sheet_names: list[str] = []
    for sname in wb.sheetnames:
        sheet = wb[sname]
        if sheet.max_row is None or sheet.max_row < 2:
            continue
        first_row = list(sheet.iter_rows(min_row=1, max_row=1))
        if not first_row:
            continue
        headers = [
            str(cell.value).strip() if cell.value else f"col_{i}"
            for i, cell in enumerate(first_row[0])
        ]
        sheet_has_data = False
        for row_cells in sheet.iter_rows(min_row=2):
            row_dict: dict[str, Any] = {"_sheet": sname}
            has_data = False
            for i, cell in enumerate(row_cells):
                if i < len(headers):
                    val = cell.value
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()
                    row_dict[headers[i]] = val
                    if val is not None:
                        has_data = True
            if has_data:
                all_rows.append(row_dict)
                sheet_has_data = True
        if sheet_has_data:
            sheet_names.append(sname)
    return all_rows, ", ".join(sheet_names) if sheet_names else "empty"


def _extract_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [r for r in reader if any(row.values())]


def _preclassify_rows(rows: list[dict[str, Any]]) -> Optional[str]:
    if not rows:
        return None
    date_cols = [k for k in rows[0].keys() if "date" in k.lower()]
    has_actual_dates = False
    for row in rows[:20]:
        for col in date_cols:
            val = row.get(col)
            if val and (isinstance(val, (datetime, date)) or (isinstance(val, str) and "-" in val and len(val) >= 8)):
                has_actual_dates = True
                break
        if has_actual_dates:
            break
    if has_actual_dates:
        return "session_import"
    week_cols = [k for k in rows[0].keys() if "week" in k.lower()]
    if week_cols:
        return "template"
    load_cols = [k for k in rows[0].keys() if any(x in k.lower() for x in ["rpe", "percentage", "%", "target"])]
    if load_cols and not has_actual_dates:
        return "template"
    return None


async def _dispatch(args):
    from import_parse_ai import generate_import_parse_report
    base64_content = args["base64_content"]
    filename = args["filename"]
    file_bytes = base64.b64decode(base64_content)
    fhash = _file_hash(file_bytes)

    if filename.lower().endswith(".xlsx"):
        rows, _sheet_name = _extract_xlsx(file_bytes)
    else:
        rows = _extract_csv(file_bytes)

    classification = _preclassify_rows(rows) or "session_import"

    try:
        program = await _get_program_store().get_program()
        current_maxes = program.get("current_maxes", {})
        current_weeks = len(set(s.get("week_number") for s in program.get("sessions", []) if s.get("week_number")))
    except Exception:
        current_maxes = {}
        current_weeks = 0

    athlete_context = {
        "current_maxes": current_maxes,
        "current_program_weeks": current_weeks,
    }

    parse_result = await generate_import_parse_report(
        file_content=json.dumps(rows[:100], indent=2, default=str),
        file_name=filename,
        classification=classification,
        athlete_context=athlete_context,
    )

    glossary_store = _get_glossary_store()
    await glossary_store.get_glossary()
    unique_names = list(set(
        ex.get("name")
        for sess in parse_result.get("sessions", [])
        for ex in sess.get("exercises", [])
        if ex.get("name")
    ))
    resolved: dict[str, str] = {}
    for name in unique_names:
        gid = await glossary_store.fuzzy_resolve(name, threshold=0.92)
        if gid:
            resolved[name] = gid
    for sess in parse_result.get("sessions", []):
        for ex in sess.get("exercises", []):
            if ex.get("name") in resolved:
                ex["glossary_id"] = resolved[ex["name"]]

    import_id = await _get_import_store().stage_import({
        "import_type": "template" if classification == "template" else "session_import",
        "source_filename": filename,
        "source_file_hash": fhash,
        "ai_parse_result": parse_result,
    })

    return {
        "import_id": import_id,
        "classification": classification,
        "warnings": parse_result.get("warnings", []),
        "parse_notes": parse_result.get("parse_notes", ""),
    }


def import_parse_file(args):
    return _run_async(_dispatch(args))

